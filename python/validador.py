from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import openpyxl
import time
   

driver = webdriver.Chrome()
driver.get("http://localhost/erasmus/interfaz/acceso/registro.html")
elem=driver.find_elements(By.TAG_NAME,"input")
button=driver.find_elements(By.TAG_NAME,"button")[0]

dataframe = openpyxl.load_workbook("/Applications/XAMPP/xamppfiles/htdocs/erasmus/python/tests.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active

def comprueba(vector,respuestas):
    # Empieza en la A18 para no escribir encima de los test
    contador=18 
    for i in range(len(vector)):
        if elem[i].is_displayed() :
            elem[i].send_keys(vector[i])
    button.click()
    time.sleep(1)
    for i in range(len(vector)):
        esInvalido = "error-input" in elem[i].get_attribute('class')
        if bool(respuestas[i]) == esInvalido and elem[i].is_displayed(): 
            print('Se ha detectado un error de validacion en el campo ' + elem[i].get_attribute('name'))
            dataframe1['A'+str(contador)] = 'Se ha detectado un error de validacion en el campo ' + elem[i].get_attribute('name')
            contador = contador + 1
            dataframe.save("/Applications/XAMPP/xamppfiles/htdocs/erasmus/python/tests.xlsx")

 
# Empezamos en 1 porque la fila cero es de headers
for row in range(1, dataframe1.max_row):
    vector=[]
    respuestas=[]
    columna=0
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        if (columna<14):
            if  ""+str(col[row].value)!="None":
                vector.append(""+str(col[row].value))
            else:
                vector.append("")
            columna=columna+1
        else:
            respuestas.append(str(col[row].value))
            columna=columna+1
    print(vector)
    print(respuestas)
    driver.execute_script("document.forms[0].reset()")
    comprueba(vector,respuestas)
    time.sleep(2)
